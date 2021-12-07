from numpy.core.defchararray import split
from numpy.core.fromnumeric import sort
from openpyxl import load_workbook
import pandas as pd
from preprocessing_data import *
from train import *

def main():
    #read data from .mat file
    boy_data = import_data('Emotrans1_Boy_data_preprocessed_42.mat')
    girl_data = import_data('Emotrans1_girl_data_preprocessed_42.mat')
    
    total_data = np.array(np.concatenate([boy_data,girl_data]))
    
    total_data = total_data[0:40]
    
    sleep_good = [0,1,38,5,17,20,21,23,24,28,37]
    sleep_bad = [2,3,9,27,29,32,7,18,12,4,14]
    
    good_data = total_data[sleep_good]
    bad_data = total_data[sleep_bad]
    
    group1_l1, group1_logistic, group1_random= generate_neutral_vs_negative(good_data,strength=[0.01,0.1,1])
    group2_l1, group2_logistic, group2_random= generate_neutral_vs_negative(bad_data,strength=[0.01,0.1,1])

    column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest"] 
    list = [group1_l1, group1_logistic, group1_random, group2_l1, group2_logistic, group2_random]
    startChar = 'A'
    wb = load_workbook('sleep_neutral_preprocess.xlsx')
    wb.create_sheet("page1")
    
    for i in range(len(list)):
        df_new = pd.DataFrame({column[i]:list[i]})
        ws = wb["page1"]
        for ind, row in df_new.iterrows():
            increase = 2
            cell =  (startChar + '%d')  % (ind + increase)
            ws[cell] = row[0]
        startChar = chr(int(ord(startChar)+1))
        if (i + 1 )%3 == 0:
            startChar = chr(int(ord(startChar)+1))
            
    boy_data = import_data('Emotrans1_Boy_data_raw.mat')
    girl_data = import_data('Emotrans1_girl_data_raw_update.mat')
    
    total_data = np.array(np.concatenate([boy_data,girl_data]))
    
    total_data = total_data[0:40]
    
    sleep_good = [0,1,38,5,17,20,21,23,24,28,37]
    sleep_bad = [2,3,9,27,29,32,7,18,12,4,14]
    
    good_data = total_data[sleep_good]
    bad_data = total_data[sleep_bad]
    
    group1_l1, group1_logistic, group1_random= generate_neutral_vs_negative(good_data,strength=[0.01,0.1,1])
    group2_l1, group2_logistic, group2_random= generate_neutral_vs_negative(bad_data,strength=[0.01,0.1,1])

    column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest"] 
    list = [group1_l1, group1_logistic, group1_random, group2_l1, group2_logistic, group2_random]
    startChar = 'A'
    wb.create_sheet("page2")
    
    for i in range(len(list)):
        df_new = pd.DataFrame({column[i]:list[i]})
        ws = wb["page2"]
        for ind, row in df_new.iterrows():
            increase = 2
            cell =  (startChar + '%d')  % (ind + increase)
            ws[cell] = row[0]
        startChar = chr(int(ord(startChar)+1))
        if (i + 1 )%3 == 0:
            startChar = chr(int(ord(startChar)+1))
            
    wb.save("sleep_neutral_preprocess.xlsx")
if __name__ == "__main__":
    main()
    
    
    