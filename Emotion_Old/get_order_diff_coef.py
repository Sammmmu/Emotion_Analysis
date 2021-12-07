from numpy.core.defchararray import split
from numpy.core.fromnumeric import sort
from openpyxl import load_workbook
import pandas as pd
from preprocessing_data import *
from train import *
from matplotlib import pyplot as plt
import mne
from mpl_toolkits.axes_grid1 import make_axes_locatable

def main():
    #read data from .mat file
    boy_data = import_data('Emotrans1_Boy_data_raw.mat')
    girl_data = import_data('Emotrans1_girl_data_raw_update.mat')
    total_data = np.array(np.concatenate([boy_data,girl_data]))
    # wb = load_workbook('order_diff_coeff_raw.xlsx')
    total_data = total_data[0:40]
    
    stress =  [22, 10, 19, 17, 23, 14, 9, 18, 23, 10, 8, 13, 12, 20, 21, 13, 0, 28, 13, 7, 18, 13, 15, 11, 13, 25, 7, 28, 14, 8, 13, 10, 12, 11, 16, 20, 25, 16, 31, 9]
    split_result = split_group_stress(total_data,stress,3)
    
    group1_data = split_result[0][0]
    group2_data = split_result[0][1]
    group3_data = split_result[0][2]

    
    group1_l1, group1_logistic, group1_random,group1_coeff= generate_diff(group1_data,clf = "elasticnet",strength=[0.01,0.1,1,10],data_type=2)
    group2_l1, group2_logistic, group2_random,group2_coeff= generate_diff(group2_data,clf = "elasticnet",strength=[0.01,0.1,1,10],data_type=2)
    group3_l1, group3_logistic, group3_random,group3_coeff= generate_diff(group3_data,clf = "elasticnet",strength=[0.01,0.1,1,10],data_type=2)
    # total_l1,total_logistic,total_random,total_coeff = generate_diff(total_data,strength=[0.01,0.1,1,10],data_type=2)

 

    coeff1_first_result = [[],[],[],[],[],[],[],[]]
    coeff2_first_result = [[],[],[],[],[],[],[],[]]
    coeff3_first_result = [[],[],[],[],[],[],[],[]]
    total_result = [[],[],[],[],[],[],[],[]]
    
    y_label = []
    for i in range(128):
        for j in range(8):
            coeff1_first_result[j].append(group1_coeff[i*8+j])
            coeff2_first_result[j].append(group2_coeff[i*8+j])
            coeff3_first_result[j].append(group3_coeff[i*8+j])
            # total_result[j].append(total_coeff[i*8+j])
        
        index1 = int(i/32)
        index2 = int(i%32)+1
        if index1 == 0:
            pre = 'A'
        elif index1 == 1:
            pre = 'B'
        elif index1 == 2:
            pre = 'C'
        else:
            pre = 'D'
        y_label.append(pre+str(index2))
    
    coeff1_first_result = np.transpose(np.array(coeff1_first_result))
    fig = plt.figure(figsize=(8,20))
    plt.tick_params(axis='y', labelsize=7)
    ax = fig.add_subplot(111)
    ax.set_yticks(range(128))
    ax.set_yticklabels(y_label)
    ax.set_xticks(range(8))
    ax.set_xticklabels(['3 8','8 10','10 13','8 13','13 20','20 30','13 30', '30 45'])
    im = ax.imshow(coeff1_first_result, cmap='YlGnBu', interpolation='nearest', aspect='auto')
    plt.title('low stress feature selection')
    plt.xlabel('Frequency bands')
    plt.ylabel('Channels')
    plt.colorbar(im)
    plt.savefig('low stress feature selection.png')

    coeff2_first_result = np.transpose(np.array(coeff2_first_result))
    fig = plt.figure(figsize=(8,20))
    plt.tick_params(axis='y', labelsize=7)
    ax = fig.add_subplot(111)
    ax.set_yticks(range(128))
    ax.set_yticklabels(y_label)
    ax.set_xticks(range(8))
    ax.set_xticklabels(['3 8','8 10','10 13','8 13','13 20','20 30','13 30', '30 45'])
    im = ax.imshow(coeff2_first_result, cmap='YlGnBu', interpolation='nearest', aspect='auto')
    plt.title('medium stress feature selection')
    plt.xlabel('Frequency bands')
    plt.ylabel('Channels')
    plt.colorbar(im)
    plt.savefig('medium stress feature selection.png')

    coeff3_first_result = np.transpose(np.array(coeff3_first_result))
    fig = plt.figure(figsize=(8,20))
    plt.tick_params(axis='y', labelsize=7)
    ax = fig.add_subplot(111)
    ax.set_yticks(range(128))
    ax.set_yticklabels(y_label)
    ax.set_xticks(range(8))
    ax.set_xticklabels(['3 8','8 10','10 13','8 13','13 20','20 30','13 30', '30 45'])
    im = ax.imshow(coeff3_first_result, cmap='YlGnBu', interpolation='nearest', aspect='auto')
    plt.title('High stress feature selection')
    plt.xlabel('Frequency bands')
    plt.ylabel('Channels')
    plt.colorbar(im)
    plt.savefig('High stress feature selection.png')

    # index = np.array(sorted(range(len(total_result)), key=lambda k: total_result[k]))
    # print(index)
    # total_result = np.transpose(np.array(sorted(total_result)[index]))
    # fig = plt.figure(figsize=(8,10))
    # plt.tick_params(axis='y', labelsize=7)
    # ax = fig.add_subplot(111)
    # ax.set_yticks(range(11))
    # ax.set_yticklabels(y_label[index])
    # ax.set_xticks(range(9))
    # ax.set_xticklabels(['0','3 8','8 10','10 13','8 13','13 20','20 30','13 30', '30 45'])
    # im = ax.imshow(total_result, cmap='YlGnBu', interpolation='nearest', aspect='auto')
    # plt.title('Total data feature selection')
    # plt.xlabel('Frequency bands')
    # plt.ylabel('Channels')
    # plt.colorbar(im)
    # plt.savefig('Total data feature selection.png')
    
    # biosemi_montage1 = mne.channels.make_standard_montage('biosemi128')
    # n_channels = len(biosemi_montage1.ch_names) 
    # fake_info = mne.create_info(ch_names=biosemi_montage1.ch_names, sfreq=250.,
    #                             ch_types='eeg')
    # fake_evoked = mne.EvokedArray(coeff1_first_result, fake_info)
    # fake_evoked.set_montage(biosemi_montage1)
    # # fig, ax = plt.subplots(ncols=3, figsize=(16, 8), gridspec_kw=dict(top=0.9),
    # #                     sharex=True, sharey=True)
    # # fake_evoked.plot_topomap(0,show_names=True, colorbar=True,size=2, res=128, title='Auditory response')
    # im,cn = mne.viz.plot_topomap(fake_evoked.data[:, 0],fake_evoked.info,names=biosemi_montage1.ch_names,show_names=True,show=False)
    # plt.title("Low Stress Feature Selection 0.5-4.5s abs")
    # clb = plt.colorbar(im)
    # clb.ax.set_title('Feature Weight',fontsize=8)
    # plt.savefig("Low Stress Feature Selection  0.5-4.5s abs.jpg")

    
    
    # biosemi_montage2 = mne.channels.make_standard_montage('biosemi128')
    # n_channels = len(biosemi_montage2.ch_names) 
    # fake_info = mne.create_info(ch_names=biosemi_montage2.ch_names, sfreq=250.,
    #                             ch_types='eeg')
    # fake_evoked = mne.EvokedArray(coeff2_first_result, fake_info)
    # fake_evoked.set_montage(biosemi_montage2)
    # im,cn =mne.viz.plot_topomap(fake_evoked.data[:, 0],fake_evoked.info,names=biosemi_montage2.ch_names,show_names=True,show=False)
    # plt.title("Mid Stress Feature Selection  0.5-4.5s abs")
    # clb = plt.colorbar(im)
    # clb.ax.set_title('Feature Weight',fontsize=8)
    # plt.savefig("Mid Stress Feature Selection  0.5-4.5s abs.jpg")
    # plt.show()
    # biosemi_montage3 = mne.channels.make_standard_montage('biosemi128')
    # n_channels = len(biosemi_montage3.ch_names) 
    # fake_info = mne.create_info(ch_names=biosemi_montage3.ch_names, sfreq=250.,
    #                             ch_types='eeg')
    # fake_evoked = mne.EvokedArray(coeff3_first_result, fake_info)
    # fake_evoked.set_montage(biosemi_montage3)
    # im,cn = mne.viz.plot_topomap(fake_evoked.data[:, 0],fake_evoked.info,names=biosemi_montage3.ch_names,show_names=True,show=False)
    # plt.title("High Stress Feature Selection 0.5-4.5s abs")
    # clb = plt.colorbar(im)
    # clb.ax.set_title('Feature Weight',fontsize=8)
    # plt.savefig("High Stress Feature Selection 0.5-4.5s abs.jpg")
    # plt.show()
    # group_l1 = np.concatenate([group1_l1,group2_l1,group3_l1])
    # group_logistic = np.concatenate([group1_logistic,group2_logistic,group3_logistic])
    # group_random = np.concatenate([group1_random,group2_random,group3_random])
    # group_coeff = np.concatenate([group1_coeff, group2_coeff,group3_coeff])
    # group_l1 = re_rank_result(group_l1,sorted(stress))
    # group_logistic  = re_rank_result(group_logistic ,sorted(stress))
    # group_random = re_rank_result(group_random,sorted(stress ))
    # column = ["L1","Logistic","Random_Forest","coeff"] 
    # list = [group_l1, group_logistic, group_random,group_coeff]
    # startChar = 'A'
    # wb.create_sheet("page3")
    # for i in range(len(list)):
    #     df_new = pd.DataFrame({column[i]:list[i]})
    #     ws = wb["page3"]
    #     for ind, row in df_new.iterrows():
    #         increase = 2
    #         cell =  (startChar + '%d')  % (ind + increase)
    #         ws[cell] = row[0]
    #     startChar = chr(int(ord(startChar)+1))
    #     if (i + 1 )%3 == 0:
    #         startChar = chr(int(ord(startChar)+1))
    # wb.save("order_diff_coeff_raw.xlsx")
    plt.show()
if __name__ == "__main__":
    main()
    
    
    