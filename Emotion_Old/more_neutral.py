from numpy.core.defchararray import split
from numpy.core.fromnumeric import sort
from openpyxl import load_workbook
import pandas as pd
from preprocessing_data import *
from train import *

def main():
    #read data from .mat file
    boy_data = import_data('Emotrans1_Boy_data_preprocessed_42.mat')
    girl_data = import_data('Emotrans1_girl_data_preprocessed_42.mat')
    
    total_data = np.array(np.concatenate([boy_data,girl_data]))
    
    neutral_index = [12,10,39,41,36,38,5,23,24,28,37,8,16,22,30,33,2,3,9,27,29,32,18,11]
    negative_index = [4,14,0,20,21,6,15,19,25,26,31,34,7]

    more_neutral = total_data[neutral_index]
    more_negative = total_data[negative_index]
    
    group1_l1, group1_logistic, group1_random= generate_diff(more_neutral,strength=[0.01,0.1,1,10])
    group2_l1, group2_logistic, group2_random= generate_diff(more_negative,strength=[0.01,0.1,1,10])

    column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest"] 
    list = [group1_l1, group1_logistic, group1_random, group2_l1, group2_logistic, group2_random]
    startChar = 'A'
    wb = load_workbook('more_neutral.xlsx')
    wb.create_sheet("page1")
    
    for i in range(len(list)):
        df_new = pd.DataFrame({column[i]:list[i]})
        ws = wb["page1"]
        for ind, row in df_new.iterrows():
            increase = 2
            cell =  (startChar + '%d')  % (ind + increase)
            ws[cell] = row[0]
        startChar = chr(int(ord(startChar)+1))
        if (i + 1 )%3 == 0:
            startChar = chr(int(ord(startChar)+1))
    
    boy_data = import_data('Emotrans1_Boy_data_raw.mat')
    girl_data = import_data('Emotrans1_girl_data_raw_update.mat')
    
    total_data = np.array(np.concatenate([boy_data,girl_data]))
    
    neutral_index = [12,10,39,41,36,38,5,23,24,28,37,8,16,22,30,33,2,3,9,27,29,32,18,11]
    negative_index = [4,14,0,20,21,6,15,19,25,26,31,34,7]
    
    more_neutral = total_data[neutral_index]
    more_negative = total_data[negative_index]
    
    group1_l1, group1_logistic, group1_random= generate_diff(more_neutral,strength=[0.01,0.1,1,10])
    group2_l1, group2_logistic, group2_random= generate_diff(more_negative,strength=[0.01,0.1,1,10])

    column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest"] 
    list = [group1_l1, group1_logistic, group1_random, group2_l1, group2_logistic, group2_random]
    startChar = 'A'
    wb.create_sheet("page2")
    
    for i in range(len(list)):
        df_new = pd.DataFrame({column[i]:list[i]})
        ws = wb["page2"]
        for ind, row in df_new.iterrows():
            increase = 2
            cell =  (startChar + '%d')  % (ind + increase)
            ws[cell] = row[0]
        startChar = chr(int(ord(startChar)+1))
        if (i + 1 )%3 == 0:
            startChar = chr(int(ord(startChar)+1))
            
    wb.save("more_neutral.xlsx")
if __name__ == "__main__":
    main()
    
    
    