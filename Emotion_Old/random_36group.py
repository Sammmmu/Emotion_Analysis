from numpy.core.defchararray import split
from numpy.core.fromnumeric import sort
from openpyxl import load_workbook
import pandas as pd
from preprocessing_data import *
from train import *

def main():
    
    boy_data = import_data('Emotrans1_Boy_data_preprocessed_42.mat')
    girl_data = import_data('Emotrans1_girl_data_preprocessed_42.mat')
    total_data = np.array(np.concatenate([boy_data,girl_data]))
    
    group1_l1_result = []
    group1_logistic_result = []
    group1_random_result = []

    group2_l1_result = []
    group2_logistic_result = []
    group2_random_result = []

    group3_l1_result = []
    group3_logistic_result = []
    group3_random_result = []
    
    wb = load_workbook('random_result_raw_10.xlsx')
    
    personal_average_l1 = []
    personal_average_logistic = []
    personal_average_random = []
    
    for i in range(36):
        personal_average_l1.append([])
        personal_average_logistic.append([])
        personal_average_random.append([])


    total_data = np.delete(total_data,[7,12,20,32,40,41])
    for j in range(1):
        stress =  [22, 10, 19, 17, 23, 14, 9, 18, 23, 10, 8, 13, 12, 20, 21, 13, 0, 28, 13, 7, 18, 13, 15, 11, 13, 25, 7, 28, 14, 8, 13, 10, 12, 11, 16, 20, 25, 16, 31, 9]
        stress = np.delete(stress,[7,12,20,32])
        # np.random.shuffle(stress)
        split_result = split_group_stress(total_data,stress,3)
        group1_data = split_result[0][0]
        group2_data = split_result[0][1]
        group3_data = split_result[0][2]


        group1_l1, group1_logistic, group1_random= generate_diff(group1_data,strength=[0.01,0.1,1,10])
        group2_l1, group2_logistic, group2_random= generate_diff(group2_data,strength=[0.01,0.1,1,10])
        group3_l1, group3_logistic, group3_random= generate_diff(group3_data,strength=[0.01,0.1,1,10])
        

        # group1_l1_result.append(sum(group1_l1)/len(group1_l1))
        # group1_logistic_result.append(sum(group1_logistic)/len(group1_logistic))
        # group1_random_result.append(sum(group1_random)/len(group1_random))
        
        # group2_l1_result.append(sum(group2_l1)/len(group2_l1))
        # group2_logistic_result.append(sum(group2_logistic)/len(group2_logistic))
        # group2_random_result.append(sum(group2_random)/len(group2_random))
        
        # group3_l1_result.append(sum(group3_l1)/len(group3_l1))
        # group3_logistic_result.append(sum(group3_logistic)/len(group3_logistic))
        # group3_random_result.append(sum(group3_random)/len(group3_random))
        
        # group_l1 = np.concatenate([group1_l1,group2_l1,group2_l1])
        # group_logistic = np.concatenate([group1_logistic, group2_logistic, group3_logistic])
        # group_random = np.concatenate([group1_random, group2_random, group3_random])
        
        # group_l1 = re_rank_result(group_l1, stress)
        # group_logistic = re_rank_result(group_logistic, stress)
        # group_random = re_rank_result(group_random, stress)
        
        list = [group1_l1, group1_logistic, group1_random,group2_l1, group2_logistic, group2_random,group3_l1, group3_logistic, group3_random]
        
        # for i in range(36):
        #     personal_average_l1[i].append(group_l1[i])
        #     personal_average_logistic[i].append(group_logistic[i])
        #     personal_average_random[i].append(group_random[i])
        
        column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest","L1","Logistic","Random_Forest"]
        startChar = 'A'
        wb.create_sheet("page"+str(j))
        for i in range(len(list)):
            df_new = pd.DataFrame({column[i]:list[i]})
            ws = wb["page"+str(j)]
            for ind, row in df_new.iterrows():
                increase = 2
                cell =  (startChar + '%d')  % (ind + increase)
                ws[cell] = row[0]
            startChar = chr(int(ord(startChar)+1))
            if (i + 1 )%3 == 0:
                startChar = chr(int(ord(startChar)+1))

    # group = [group1_l1_result,group1_logistic_result,group1_random_result,group2_l1_result,group2_logistic_result,group2_random_result,group3_l1_result,group3_logistic_result,group3_random_result]
    
    # column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest","L1","Logistic","Random_Forest",]
    
    # wb.create_sheet("page101")
    
    # for k in range(len(group)):
    #     df_new = pd.DataFrame({column[k]:group[k]})
    #     ws = wb["page101"]
    #     for ind, row in df_new.iterrows():
    #         increase = 2
    #         cell =  (startChar + '%d')  % (ind + increase)
    #         ws[cell] = row[0]
    #     startChar = chr(int(ord(startChar)+1))
    #     if (k + 1 )%3 == 0:
    #         startChar = chr(int(ord(startChar)+1))
    # personal_average_1 = []
    # personal_average_2 = []
    # personal_average_3 = []
    # for i in range(36):
    #     personal_average_1.append(sum(personal_average_l1[i])/len(personal_average_l1[i]))
    #     personal_average_2.append(sum(personal_average_logistic[i])/len(personal_average_logistic[i]))
    #     personal_average_3.append(sum(personal_average_random[i])/len(personal_average_random[i]))
    
    # wb.create_sheet("page102")
    # group_2 = [personal_average_1,personal_average_2,personal_average_3]
    # column = ["L1","Logistic","Random_Forest","L1"]
    # for k in range(len(group_2)):
    #     df_new = pd.DataFrame({column[k]:group_2[k]})
    #     ws = wb["page102"]
    #     for ind, row in df_new.iterrows():
    #         increase = 2
    #         cell =  (startChar + '%d')  % (ind + increase)
    #         ws[cell] = row[0]
    #     startChar = chr(int(ord(startChar)+1))
    #     if (k + 1 )%3 == 0:
    #         startChar = chr(int(ord(startChar)+1))
    
    # personal_average_1 = []
    # personal_average_2 = []
    # personal_average_3 = []
    # for i in range(36):
    #     personal_average_1.append(np.std(np.array(personal_average_l1[i])))
    #     personal_average_2.append(np.std(np.array(personal_average_logistic[i])))
    #     personal_average_3.append(np.std(np.array(personal_average_random[i])))
    
    # wb.create_sheet("page103")
    # group_2 = [personal_average_1,personal_average_2,personal_average_3]
    # column = ["L1","Logistic","Random_Forest","L1"]
    # for k in range(len(group_2)):
    #     df_new = pd.DataFrame({column[k]:group_2[k]})
    #     ws = wb["page103"]
    #     for ind, row in df_new.iterrows():
    #         increase = 2
    #         cell =  (startChar + '%d')  % (ind + increase)
    #         ws[cell] = row[0]
    #     startChar = chr(int(ord(startChar)+1))
    #     if (k + 1 )%3 == 0:
    #         startChar = chr(int(ord(startChar)+1))

    wb.save('random_result_raw_10.xlsx')
    
  #read data from .mat file
    # boy_data = import_data('Emotrans1_Boy_data_preprocessed_42.mat')
    # girl_data = import_data('Emotrans1_girl_data_preprocessed_42.mat')
    # total_data = np.array(np.concatenate([boy_data,girl_data]))
    
    # group1_l1_result = []
    # group1_logistic_result = []
    # group1_random_result = []

    # group2_l1_result = []
    # group2_logistic_result = []
    # group2_random_result = []

    # group3_l1_result = []
    # group3_logistic_result = []
    # group3_random_result = []
    
    # wb = load_workbook('random_result_raw_8.xlsx')
    
    # personal_average_l1 = []
    # personal_average_logistic = []
    # personal_average_random = []
    
    # for i in range(36):
    #     personal_average_l1.append([])
        
        
        
    #     personal_average_logistic.append([])
    #     personal_average_random.append([])


    # total_data = np.delete(total_data,[7,12,20,32])
    # for j in range(100):
    #     stress =  [22, 10, 19, 17, 23, 14, 9, 18, 23, 10, 8, 13, 12, 20, 21, 13, 0, 28, 13, 7, 18, 13, 15, 11, 13, 25, 7, 28, 14, 8, 13, 10, 12, 11, 16, 20, 25, 16, 31, 9]
    #     stress = np.delete(stress,[7,12,20,32])
    #     np.random.shuffle(stress)
    #     split_result = split_group_stress(total_data,stress,3)
    #     group1_data = split_result[0][0]
    #     group1_index = split_result[1][0]
    #     group2_data = split_result[0][1]
    #     group2_index = split_result[1][1]
    #     group3_data = split_result[0][2]
    #     group3_index = split_result[1][2]

    #     group1_l1, group1_logistic, group1_random= generate_diff(group1_data,strength=[0.01,0.1,1,10])
    #     group2_l1, group2_logistic, group2_random= generate_diff(group2_data,strength=[0.01,0.1,1,10])
    #     group3_l1, group3_logistic, group3_random= generate_diff(group3_data,strength=[0.01,0.1,1,10])
        

    #     group1_l1_result.append(sum(group1_l1)/len(group1_l1))
    #     group1_logistic_result.append(sum(group1_logistic)/len(group1_logistic))
    #     group1_random_result.append(sum(group1_random)/len(group1_random))
        
    #     group2_l1_result.append(sum(group2_l1)/len(group2_l1))
    #     group2_logistic_result.append(sum(group2_logistic)/len(group2_logistic))
    #     group2_random_result.append(sum(group2_random)/len(group2_random))
        
    #     group3_l1_result.append(sum(group3_l1)/len(group3_l1))
    #     group3_logistic_result.append(sum(group3_logistic)/len(group3_logistic))
    #     group3_random_result.append(sum(group3_random)/len(group3_random))
        
    #     group_l1 = np.concatenate([group1_l1,group2_l1,group2_l1])
    #     group_logistic = np.concatenate([group1_logistic, group2_logistic, group3_logistic])
    #     group_random = np.concatenate([group1_random, group2_random, group3_random])
        
    #     group_l1 = re_rank_result(group_l1, stress)
    #     group_logistic = re_rank_result(group_logistic, stress)
    #     group_random = re_rank_result(group_random, stress)
        
    #     list = [group_l1, group_logistic, group_random]
        
    #     for i in range(36):
    #         personal_average_l1[i].append(group_l1[i])
    #         personal_average_logistic[i].append(group_logistic[i])
    #         personal_average_random[i].append(group_random[i])
        
    #     column = ["L1","Logistic","Random_Forest","L1","Logistic"]
    #     startChar = 'A'
    #     wb.create_sheet("page"+str(j))
    #     for i in range(len(list)):
    #         df_new = pd.DataFrame({column[i]:list[i]})
    #         ws = wb["page"+str(j)]
    #         for ind, row in df_new.iterrows():
    #             increase = 2
    #             cell =  (startChar + '%d')  % (ind + increase)
    #             ws[cell] = row[0]
    #         startChar = chr(int(ord(startChar)+1))
    #         if (i + 1 )%3 == 0:
    #             startChar = chr(int(ord(startChar)+1))

    # group = [group1_l1_result,group1_logistic_result,group1_random_result,group2_l1_result,group2_logistic_result,group2_random_result,group3_l1_result,group3_logistic_result,group3_random_result]
    
    # column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest","L1","Logistic","Random_Forest",]
    
    # wb.create_sheet("page101")
    
    # for k in range(len(group)):
    #     df_new = pd.DataFrame({column[k]:group[k]})
    #     ws = wb["page101"]
    #     for ind, row in df_new.iterrows():
    #         increase = 2
    #         cell =  (startChar + '%d')  % (ind + increase)
    #         ws[cell] = row[0]
    #     startChar = chr(int(ord(startChar)+1))
    #     if (k + 1 )%3 == 0:
    #         startChar = chr(int(ord(startChar)+1))
    # personal_average_1 = []
    # personal_average_2 = []
    # personal_average_3 = []
    # for i in range(36):
    #     personal_average_1.append(sum(personal_average_l1[i])/len(personal_average_l1[i]))
    #     personal_average_2.append(sum(personal_average_logistic[i])/len(personal_average_logistic[i]))
    #     personal_average_3.append(sum(personal_average_random[i])/len(personal_average_random[i]))
    
    # wb.create_sheet("page102")
    # group_2 = [personal_average_1,personal_average_2,personal_average_3]
    # column = ["L1","Logistic","Random_Forest","L1"]
    # for k in range(len(group_2)):
    #     df_new = pd.DataFrame({column[k]:group_2[k]})
    #     ws = wb["page102"]
    #     for ind, row in df_new.iterrows():
    #         increase = 2
    #         cell =  (startChar + '%d')  % (ind + increase)
    #         ws[cell] = row[0]
    #     startChar = chr(int(ord(startChar)+1))
    #     if (k + 1 )%3 == 0:
    #         startChar = chr(int(ord(startChar)+1))
    
    # personal_average_1 = []
    # personal_average_2 = []
    # personal_average_3 = []
    # for i in range(36):
    #     personal_average_1.append(np.std(np.array(personal_average_l1[i])))
    #     personal_average_2.append(np.std(np.array(personal_average_logistic[i])))
    #     personal_average_3.append(np.std(np.array(personal_average_random[i])))
    
    # wb.create_sheet("page103")
    # group_2 = [personal_average_1,personal_average_2,personal_average_3]
    # column = ["L1","Logistic","Random_Forest","L1"]
    # for k in range(len(group_2)):
    #     df_new = pd.DataFrame({column[k]:group_2[k]})
    #     ws = wb["page103"]
    #     for ind, row in df_new.iterrows():
    #         increase = 2
    #         cell =  (startChar + '%d')  % (ind + increase)
    #         ws[cell] = row[0]
    #     startChar = chr(int(ord(startChar)+1))
    #     if (k + 1 )%3 == 0:
    #         startChar = chr(int(ord(startChar)+1))

    # wb.save('random_result_raw_8.xlsx')

if __name__ == "__main__":
    main()
    