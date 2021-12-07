from numpy.core.defchararray import split
from numpy.core.fromnumeric import sort
from openpyxl import load_workbook
import pandas as pd
from preprocessing_data import *
from train import *

def main():
    #read data from .mat file

    total_data = import_data('Emotrans2_fall_raw.mat')

    wb = load_workbook('new_data_order_diff_elasticNet.xlsx')
    total_data = np.array(total_data)
    for i in range(len(total_data)):
        total_data[i] = total_data[i][0:2]

    group1_l1, group1_logistic, group1_random= generate_diff(total_data,clf = "elasticnet",strength=[0.01,0.1,1])


    list = [group1_l1, group1_logistic, group1_random] 
    column = ["L1","Logistic","Random_Forest"]
    startChar = 'A'
    wb.create_sheet("C-watch-Reappraisal")
    for i in range(len(list)):
        df_new = pd.DataFrame({column[i]:list[i]})
        ws = wb["C-watch-Reappraisal"] 
        for ind, row in df_new.iterrows():
            increase = 2
            cell =  (startChar + '%d')  % (ind + increase)
            ws[cell] = row[0]
        startChar = chr(int(ord(startChar)+1))
        if (i + 1 )%3 == 0:
            startChar = chr(int(ord(startChar)+1))


    wb.save('new_data_order_diff_elasticNet.xlsx')

if __name__ == "__main__":
    main()
    
    
    