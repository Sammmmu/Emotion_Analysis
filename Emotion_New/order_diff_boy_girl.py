from numpy.core.defchararray import split
from numpy.core.fromnumeric import sort
from openpyxl import load_workbook
import pandas as pd
from preprocessing_data import *
from train import *

def main():
    #read data from .mat file
    boy_data = import_data('Emotrans1_Boy_data_raw.mat')
    girl_data = import_data('Emotrans1_girl_data_raw_update.mat')
    total_data = np.array(np.concatenate([boy_data,girl_data]))
    
    
    wb = load_workbook('order_diff_boy_girl.xlsx')

    stress =  [22, 10, 19, 17, 23, 14, 9, 18, 23, 10, 8, 13, 12, 20, 21, 13, 0, 28, 13, 7, 18, 13, 15, 11, 13, 25, 7, 28, 14, 8, 13, 10, 12, 11, 16, 20, 25, 16, 31, 9]
    boy_index = [0,1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,19,21,22]
    girl_index = [17,18,20,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39]
    boy_data = total_data[boy_index]
    girl_data =total_data[girl_index]

    group1_l1, group1_logistic, group1_random= generate_diff(boy_data,strength=[0.01,0.1,1,10])
    group2_l1, group2_logistic, group2_random= generate_diff(girl_data,strength=[0.01,0.1,1,10])


    list = [group1_l1, group1_logistic, group1_random,group2_l1, group2_logistic, group2_random] 
    column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest"]
    startChar = 'A'
    wb.create_sheet("page1")
    for i in range(len(list)):
        df_new = pd.DataFrame({column[i]:list[i]})
        ws = wb["page1"]
        for ind, row in df_new.iterrows():
            increase = 2
            cell =  (startChar + '%d')  % (ind + increase)
            ws[cell] = row[0]
        startChar = chr(int(ord(startChar)+1))
        if (i + 1 )%3 == 0:
            startChar = chr(int(ord(startChar)+1))

    boy_data = import_data('Emotrans1_Boy_data_preprocessed_42.mat')
    girl_data = import_data('Emotrans1_girl_data_preprocessed_42.mat')
    total_data = np.array(np.concatenate([boy_data,girl_data]))
    
    boy_data = total_data[boy_index]
    girl_data =total_data[girl_index]

    group1_l1, group1_logistic, group1_random= generate_diff(boy_data,strength=[0.01,0.1,1,10])
    group2_l1, group2_logistic, group2_random= generate_diff(girl_data,strength=[0.01,0.1,1,10])


    list = [group1_l1, group1_logistic, group1_random,group2_l1, group2_logistic, group2_random] 
    column = ["L1","Logistic","Random_Forest","L1","Logistic","Random_Forest"]
    startChar = 'A'
    wb.create_sheet("page2")
    for i in range(len(list)):
        df_new = pd.DataFrame({column[i]:list[i]})
        ws = wb["page2"]
        for ind, row in df_new.iterrows():
            increase = 2
            cell =  (startChar + '%d')  % (ind + increase)
            ws[cell] = row[0]
        startChar = chr(int(ord(startChar)+1))
        if (i + 1 )%3 == 0:
            startChar = chr(int(ord(startChar)+1))

    wb.save('order_diff_boy_girl.xlsx')

if __name__ == "__main__":
    main()
    
    
    